from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Iterable, List, Dict, Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


def write_excel_report(results: Iterable[Dict[str, Any]], output_path: str | Path) -> Path:
    """Write a summarized Excel report for Appium E2E results."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "E2E Summary"

    headers = ["Test Name", "Status", "Duration (s)", "Message", "Timestamp"]
    sheet.append(headers)

    for result in results:
        sheet.append([
            result.get("test_name", "Unknown"),
            result.get("status", "unknown"),
            result.get("duration", 0),
            result.get("message", ""),
            result.get("timestamp", datetime.now().isoformat(timespec="seconds")),
        ])

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for column in sheet.columns:
        max_length = max(len(str(cell.value)) for cell in column if cell.value is not None)
        sheet.column_dimensions[column[0].column_letter].width = min(max_length + 2, 40)

    workbook.save(output_path)
    return output_path
